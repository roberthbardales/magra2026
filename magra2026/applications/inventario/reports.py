"""
applications/inventario/reports.py
Generación de reportes PDF y Excel para el módulo inventario.
Requiere: pip install reportlab openpyxl
"""
from io import BytesIO
from datetime import datetime

# ── ReportLab ──
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ── OpenPyXL ──
import openpyxl
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════
#  CONSTANTES DE ESTILO
# ══════════════════════════════════════════
AZUL       = colors.HexColor('#1a6fff')
AZUL_CLARO = colors.HexColor('#e8f0ff')
GRIS_OSC   = colors.HexColor('#1c1c1a')
GRIS_MED   = colors.HexColor('#5a5a56')
GRIS_CLR   = colors.HexColor('#f7f8fa')
VERDE      = colors.HexColor('#00a86b')
ROJO       = colors.HexColor('#e05c5c')
NARANJA    = colors.HexColor('#f0a050')
BLANCO     = colors.white

EMPRESA    = 'MAGRA'
SISTEMA    = 'Sistema ERP de Inventario'


# ══════════════════════════════════════════
#  HELPERS PDF
# ══════════════════════════════════════════
def _estilos():
    base = getSampleStyleSheet()
    estilos = {
        'titulo': ParagraphStyle(
            'titulo', fontName='Helvetica-Bold', fontSize=18,
            textColor=GRIS_OSC, spaceAfter=4, alignment=TA_LEFT
        ),
        'subtitulo': ParagraphStyle(
            'subtitulo', fontName='Helvetica', fontSize=10,
            textColor=GRIS_MED, spaceAfter=2, alignment=TA_LEFT
        ),
        'fecha': ParagraphStyle(
            'fecha', fontName='Helvetica', fontSize=8,
            textColor=GRIS_MED, spaceAfter=12, alignment=TA_RIGHT
        ),
        'celda': ParagraphStyle(
            'celda', fontName='Helvetica', fontSize=8,
            textColor=GRIS_OSC, leading=10
        ),
        'celda_bold': ParagraphStyle(
            'celda_bold', fontName='Helvetica-Bold', fontSize=8,
            textColor=GRIS_OSC, leading=10
        ),
    }
    return estilos

def _header_pdf(elements, titulo, subtitulo=''):
    e = _estilos()
    now = datetime.now().strftime('%d/%m/%Y %H:%M')
    data = [[
        Paragraph(f'<b>{EMPRESA}</b> — {titulo}', e['titulo']),
        Paragraph(now, e['fecha'])
    ]]
    t = Table(data, colWidths=['*', 4*cm])
    t.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(t)
    if subtitulo:
        elements.append(Paragraph(subtitulo, e['subtitulo']))
    elements.append(HRFlowable(width='100%', thickness=1.5, color=AZUL, spaceAfter=12))

def _tabla_estilo(col_widths, header_color=None):
    hc = header_color or AZUL
    return TableStyle([
        # Header
        ('BACKGROUND',    (0,0), (-1,0), hc),
        ('TEXTCOLOR',     (0,0), (-1,0), BLANCO),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0), 8),
        ('ALIGN',         (0,0), (-1,0), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 7),
        ('TOPPADDING',    (0,0), (-1,0), 7),
        # Filas
        ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',      (0,1), (-1,-1), 7.5),
        ('BOTTOMPADDING', (0,1), (-1,-1), 5),
        ('TOPPADDING',    (0,1), (-1,-1), 5),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [BLANCO, GRIS_CLR]),
        # Bordes
        ('GRID',          (0,0), (-1,-1), 0.3, colors.HexColor('#dde1ea')),
        ('LINEBELOW',     (0,0), (-1,0), 1, hc),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [BLANCO, GRIS_CLR]),
    ])


# ══════════════════════════════════════════
#  HELPERS EXCEL
# ══════════════════════════════════════════
def _wb_header(ws, titulo, headers, col_widths):
    # Título
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    ws['A1'] = f'{EMPRESA} — {titulo}'
    ws['A1'].font = Font(bold=True, size=14, color='1a6fff')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    # Fecha
    ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(size=9, color='5a5a56')
    ws['A2'].alignment = Alignment(horizontal='left')
    ws.row_dimensions[2].height = 16

    # Headers
    header_fill = PatternFill('solid', fgColor='1a6fff')
    header_font = Font(bold=True, color='FFFFFF', size=9)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = col_widths[col-1]
    ws.row_dimensions[3].height = 22

    # Bordes helper
    thin = Side(style='thin', color='dde1ea')
    return thin

def _fila_excel(ws, row, data, thin, alt=False):
    fill = PatternFill('solid', fgColor='f7f8fa') if alt else PatternFill('solid', fgColor='FFFFFF')
    border = Border(left=Side(style='thin', color='dde1ea'),
                    right=Side(style='thin', color='dde1ea'),
                    top=Side(style='thin', color='dde1ea'),
                    bottom=Side(style='thin', color='dde1ea'))
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = border
        cell.font = Font(size=9)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 18


# ══════════════════════════════════════════
#  1. REPORTE STOCK ACTUAL
# ══════════════════════════════════════════
def reporte_stock_pdf(productos):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    e = _estilos()
    _header_pdf(elements, 'Reporte de Stock Actual',
                f'Total productos: {len(productos)}')

    headers = ['SKU', 'Producto', 'Categoría', 'Almacén',
               'Unidad', 'Stock Actual', 'Stock Mín.', 'Stock Máx.', 'Estado']
    data = [headers]
    for p in productos:
        if p.stock_actual <= p.stock_minimo:
            estado = 'BAJO'
        elif p.stock_actual >= p.stock_maximo:
            estado = 'EXCESO'
        else:
            estado = 'OK'
        data.append([
            p.sku,
            Paragraph(p.nombre, e['celda']),
            p.categoria.nombre if p.categoria else '—',
            p.almacen.nombre if p.almacen else '—',
            p.unidad_medida.abreviatura if p.unidad_medida else '—',
            str(p.stock_actual),
            str(p.stock_minimo),
            str(p.stock_maximo),
            estado,
        ])

    col_w = [2.2*cm, 4.5*cm, 2.5*cm, 2.5*cm, 1.5*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.5*cm]
    t = Table(data, colWidths=col_w, repeatRows=1)
    estilo = _tabla_estilo(col_w)
    # Color estado
    for i, p in enumerate(productos, 1):
        if p.stock_actual <= p.stock_minimo:
            estilo.add('TEXTCOLOR', (8,i), (8,i), ROJO)
            estilo.add('FONTNAME',  (8,i), (8,i), 'Helvetica-Bold')
        elif p.stock_actual >= p.stock_maximo:
            estilo.add('TEXTCOLOR', (8,i), (8,i), NARANJA)
            estilo.add('FONTNAME',  (8,i), (8,i), 'Helvetica-Bold')
        else:
            estilo.add('TEXTCOLOR', (8,i), (8,i), VERDE)
    t.setStyle(estilo)
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def reporte_stock_excel(productos):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Stock Actual'
    headers = ['SKU','Producto','Categoría','Almacén','Unidad',
               'Stock Actual','Stock Mín.','Stock Máx.','Estado']
    col_widths = [12, 30, 16, 16, 10, 12, 12, 12, 10]
    thin = _wb_header(ws, 'Reporte de Stock Actual', headers, col_widths)
    for i, p in enumerate(productos):
        if p.stock_actual <= p.stock_minimo:
            estado = 'BAJO'
            color_estado = 'e05c5c'
        elif p.stock_actual >= p.stock_maximo:
            estado = 'EXCESO'
            color_estado = 'f0a050'
        else:
            estado = 'OK'
            color_estado = '00a86b'
        row = i + 4
        _fila_excel(ws, row, [
            p.sku, p.nombre,
            p.categoria.nombre if p.categoria else '—',
            p.almacen.nombre if p.almacen else '—',
            p.unidad_medida.abreviatura if p.unidad_medida else '—',
            p.stock_actual, p.stock_minimo, p.stock_maximo, estado
        ], thin, alt=i%2==1)
        # Color estado
        ws.cell(row=row, column=9).font = Font(bold=True, color=color_estado, size=9)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ══════════════════════════════════════════
#  2. KARDEX POR PRODUCTO
# ══════════════════════════════════════════
def reporte_kardex_pdf(producto, movimientos):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    e = _estilos()
    _header_pdf(elements, f'Kardex — {producto.nombre}',
                f'SKU: {producto.sku} | Stock actual: {producto.stock_actual}')

    headers = ['Fecha', 'Tipo', 'Documento', 'Entrada', 'Salida', 'Saldo', 'P. Unitario', 'Usuario', 'Motivo']
    data = [headers]
    saldo = 0
    TIPOS_ENTRADA = ['EC', 'DE', 'AP']
    for m in movimientos:
        es_entrada = m.tipo in TIPOS_ENTRADA
        entrada = m.cantidad if es_entrada else ''
        salida  = m.cantidad if not es_entrada else ''
        saldo  += m.cantidad if es_entrada else -m.cantidad
        data.append([
            m.fecha.strftime('%d/%m/%Y %H:%M'),
            m.get_tipo_display(),
            m.documento_referencia or '—',
            str(entrada) if entrada != '' else '—',
            str(salida)  if salida  != '' else '—',
            str(saldo),
            f"S/ {m.precio_unitario:.2f}" if m.precio_unitario else '—',
            m.usuario.full_name if m.usuario else '—',
            Paragraph(m.motivo or '—', e['celda']),
        ])

    col_w = [2.8*cm, 2.2*cm, 2.2*cm, 1.8*cm, 1.8*cm, 1.8*cm, 2*cm, 3*cm, 4.4*cm]
    t = Table(data, colWidths=col_w, repeatRows=1)
    estilo = _tabla_estilo(col_w)
    for i, m in enumerate(movimientos, 1):
        if m.tipo in TIPOS_ENTRADA:
            estilo.add('TEXTCOLOR', (3,i), (3,i), VERDE)
        else:
            estilo.add('TEXTCOLOR', (4,i), (4,i), ROJO)
    t.setStyle(estilo)
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def reporte_kardex_excel(producto, movimientos):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Kardex {producto.sku}'
    headers = ['Fecha','Tipo','Documento','Entrada','Salida','Saldo','P. Unitario','Usuario','Motivo']
    col_widths = [18, 14, 14, 10, 10, 10, 12, 22, 30]
    thin = _wb_header(ws, f'Kardex — {producto.nombre}', headers, col_widths)
    TIPOS_ENTRADA = ['EC', 'DE', 'AP']
    saldo = 0
    for i, m in enumerate(movimientos):
        es_entrada = m.tipo in TIPOS_ENTRADA
        entrada = m.cantidad if es_entrada else None
        salida  = m.cantidad if not es_entrada else None
        saldo  += m.cantidad if es_entrada else -m.cantidad
        row = i + 4
        _fila_excel(ws, row, [
            m.fecha.strftime('%d/%m/%Y %H:%M'),
            m.get_tipo_display(),
            m.documento_referencia or '—',
            entrada, salida, saldo,
            float(m.precio_unitario) if m.precio_unitario else None,
            m.usuario.full_name if m.usuario else '—',
            m.motivo or '—',
        ], thin, alt=i%2==1)
        if entrada:
            ws.cell(row=row, column=4).font = Font(bold=True, color='00a86b', size=9)
        if salida:
            ws.cell(row=row, column=5).font = Font(bold=True, color='e05c5c', size=9)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ══════════════════════════════════════════
#  3. MOVIMIENTOS POR FECHA
# ══════════════════════════════════════════
def reporte_movimientos_pdf(movimientos, fecha_inicio=None, fecha_fin=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    e = _estilos()
    rango = ''
    if fecha_inicio and fecha_fin:
        rango = f'Período: {fecha_inicio} al {fecha_fin}'
    _header_pdf(elements, 'Reporte de Movimientos', rango)

    headers = ['Fecha', 'Producto', 'SKU', 'Tipo', 'Cantidad', 'P. Unitario', 'Documento', 'Usuario']
    data = [headers]
    TIPOS_ENTRADA = ['EC', 'DE', 'AP']
    for m in movimientos:
        data.append([
            m.fecha.strftime('%d/%m/%Y'),
            Paragraph(m.producto.nombre, e['celda']),
            m.producto.sku,
            m.get_tipo_display(),
            str(m.cantidad),
            f"S/ {m.precio_unitario:.2f}" if m.precio_unitario else '—',
            m.documento_referencia or '—',
            m.usuario.full_name if m.usuario else '—',
        ])

    col_w = [2.2*cm, 4.5*cm, 2*cm, 2.5*cm, 2*cm, 2.2*cm, 2.5*cm, 3.5*cm]
    t = Table(data, colWidths=col_w, repeatRows=1)
    estilo = _tabla_estilo(col_w)
    for i, m in enumerate(movimientos, 1):
        if m.tipo in TIPOS_ENTRADA:
            estilo.add('TEXTCOLOR', (4,i), (4,i), VERDE)
        else:
            estilo.add('TEXTCOLOR', (4,i), (4,i), ROJO)
    t.setStyle(estilo)
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def reporte_movimientos_excel(movimientos, fecha_inicio=None, fecha_fin=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Movimientos'
    headers = ['Fecha','Producto','SKU','Tipo','Cantidad','P. Unitario','Documento','Usuario']
    col_widths = [14, 30, 12, 16, 10, 12, 16, 24]
    thin = _wb_header(ws, 'Reporte de Movimientos', headers, col_widths)
    TIPOS_ENTRADA = ['EC', 'DE', 'AP']
    for i, m in enumerate(movimientos):
        row = i + 4
        _fila_excel(ws, row, [
            m.fecha.strftime('%d/%m/%Y'),
            m.producto.nombre,
            m.producto.sku,
            m.get_tipo_display(),
            m.cantidad,
            float(m.precio_unitario) if m.precio_unitario else None,
            m.documento_referencia or '—',
            m.usuario.full_name if m.usuario else '—',
        ], thin, alt=i%2==1)
        color = '00a86b' if m.tipo in TIPOS_ENTRADA else 'e05c5c'
        ws.cell(row=row, column=5).font = Font(bold=True, color=color, size=9)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ══════════════════════════════════════════
#  4. ALERTAS ACTIVAS
# ══════════════════════════════════════════
def reporte_alertas_pdf(alertas):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    e = _estilos()
    activas = [a for a in alertas if not a.resuelta]
    _header_pdf(elements, 'Reporte de Alertas',
                f'Alertas activas: {len(activas)} de {len(alertas)} total')

    headers = ['Tipo', 'Producto', 'SKU', 'Mensaje', 'Prioridad', 'Fecha', 'Estado']
    data = [headers]
    COLORES_TIPO = {'SM': ROJO, 'SX': NARANJA, 'VE': colors.HexColor('#b482ff'), 'SI': GRIS_MED}
    for a in alertas:
        data.append([
            a.get_tipo_display(),
            Paragraph(a.producto.nombre, e['celda']),
            a.producto.sku,
            Paragraph(a.mensaje, e['celda']),
            str(a.prioridad),
            a.created.strftime('%d/%m/%Y') if hasattr(a, 'created') else '—',
            'Resuelta' if a.resuelta else 'Activa',
        ])

    col_w = [2.5*cm, 4*cm, 2*cm, 5.5*cm, 1.8*cm, 2.2*cm, 1.8*cm]
    t = Table(data, colWidths=col_w, repeatRows=1)
    estilo = _tabla_estilo(col_w, NARANJA)
    for i, a in enumerate(alertas, 1):
        color = COLORES_TIPO.get(a.tipo, GRIS_MED)
        estilo.add('TEXTCOLOR', (0,i), (0,i), color)
        estilo.add('FONTNAME',  (0,i), (0,i), 'Helvetica-Bold')
        if not a.resuelta:
            estilo.add('TEXTCOLOR', (6,i), (6,i), ROJO)
        else:
            estilo.add('TEXTCOLOR', (6,i), (6,i), VERDE)
    t.setStyle(estilo)
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def reporte_alertas_excel(alertas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Alertas'
    headers = ['Tipo','Producto','SKU','Mensaje','Prioridad','Fecha','Estado']
    col_widths = [14, 28, 12, 40, 10, 14, 12]
    thin = _wb_header(ws, 'Reporte de Alertas', headers, col_widths)
    COLORES_TIPO = {'SM':'e05c5c','SX':'f0a050','VE':'b482ff','SI':'9a9a94'}
    for i, a in enumerate(alertas):
        row = i + 4
        _fila_excel(ws, row, [
            a.get_tipo_display(),
            a.producto.nombre,
            a.producto.sku,
            a.mensaje,
            a.prioridad,
            a.created.strftime('%d/%m/%Y') if hasattr(a, 'created') else '—',
            'Resuelta' if a.resuelta else 'Activa',
        ], thin, alt=i%2==1)
        color_tipo = COLORES_TIPO.get(a.tipo, '9a9a94')
        ws.cell(row=row, column=1).font = Font(bold=True, color=color_tipo, size=9)
        color_estado = '00a86b' if a.resuelta else 'e05c5c'
        ws.cell(row=row, column=7).font = Font(bold=True, color=color_estado, size=9)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ══════════════════════════════════════════
#  5. AJUSTES DE INVENTARIO
# ══════════════════════════════════════════
def reporte_ajustes_pdf(ajustes):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    e = _estilos()
    _header_pdf(elements, 'Reporte de Ajustes de Inventario',
                f'Total ajustes: {len(ajustes)}')

    headers = ['Producto', 'SKU', 'Tipo', 'Cantidad', 'Motivo', 'Estado', 'Solicitado por', 'Aprobado por', 'Fecha']
    data = [headers]
    COLORES_ESTADO = {'P': NARANJA, 'A': VERDE, 'R': ROJO}
    for a in ajustes:
        data.append([
            Paragraph(a.producto.nombre, e['celda']),
            a.producto.sku,
            a.get_tipo_display() if hasattr(a, 'get_tipo_display') else a.tipo,
            str(a.cantidad),
            Paragraph(a.motivo or '—', e['celda']),
            a.get_estado_display(),
            a.solicitado_por.full_name if a.solicitado_por else '—',
            a.aprobado_por.full_name if a.aprobado_por else '—',
            a.created.strftime('%d/%m/%Y') if hasattr(a, 'created') else '—',
        ])

    col_w = [3.5*cm, 1.8*cm, 2*cm, 1.8*cm, 4*cm, 1.8*cm, 3*cm, 3*cm, 2*cm]
    t = Table(data, colWidths=col_w, repeatRows=1)
    estilo = _tabla_estilo(col_w, colors.HexColor('#b482ff'))
    for i, a in enumerate(ajustes, 1):
        color = COLORES_ESTADO.get(a.estado, GRIS_MED)
        estilo.add('TEXTCOLOR', (5,i), (5,i), color)
        estilo.add('FONTNAME',  (5,i), (5,i), 'Helvetica-Bold')
    t.setStyle(estilo)
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def reporte_ajustes_excel(ajustes):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ajustes'
    headers = ['Producto','SKU','Tipo','Cantidad','Motivo','Estado','Solicitado por','Aprobado por','Fecha']
    col_widths = [28, 12, 14, 10, 35, 12, 22, 22, 14]
    thin = _wb_header(ws, 'Reporte de Ajustes de Inventario', headers, col_widths)
    COLORES_ESTADO = {'P':'f0a050','A':'00a86b','R':'e05c5c'}
    for i, a in enumerate(ajustes):
        row = i + 4
        _fila_excel(ws, row, [
            a.producto.nombre,
            a.producto.sku,
            a.get_tipo_display() if hasattr(a, 'get_tipo_display') else a.tipo,
            a.cantidad,
            a.motivo or '—',
            a.get_estado_display(),
            a.solicitado_por.full_name if a.solicitado_por else '—',
            a.aprobado_por.full_name if a.aprobado_por else '—',
            a.created.strftime('%d/%m/%Y') if hasattr(a, 'created') else '—',
        ], thin, alt=i%2==1)
        color = COLORES_ESTADO.get(a.estado, '9a9a94')
        ws.cell(row=row, column=6).font = Font(bold=True, color=color, size=9)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ══════════════════════════════════════════
#  6. PROVEEDORES Y PRODUCTOS
# ══════════════════════════════════════════
def reporte_proveedores_pdf(proveedores):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    e = _estilos()
    _header_pdf(elements, 'Reporte de Proveedores',
                f'Total proveedores: {len(proveedores)}')

    headers = ['Proveedor', 'RUC', 'Contacto', 'Email', 'Teléfono', 'Productos']
    data = [headers]
    for p in proveedores:
        n_prod = p.producto_set.count() if hasattr(p, 'producto_set') else '—'
        data.append([
            Paragraph(p.nombre, e['celda_bold']),
            p.ruc or '—',
            p.contacto or '—',
            p.email or '—',
            p.telefono or '—',
            str(n_prod),
        ])

    col_w = [4.5*cm, 2.5*cm, 3*cm, 4*cm, 2.5*cm, 2*cm]
    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(_tabla_estilo(col_w, VERDE))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def reporte_proveedores_excel(proveedores):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proveedores'
    headers = ['Proveedor','RUC','Contacto','Email','Teléfono','Nº Productos']
    col_widths = [28, 14, 20, 28, 14, 14]
    thin = _wb_header(ws, 'Reporte de Proveedores', headers, col_widths)
    for i, p in enumerate(proveedores):
        n_prod = p.producto_set.count() if hasattr(p, 'producto_set') else '—'
        row = i + 4
        _fila_excel(ws, row, [
            p.nombre, p.ruc or '—', p.contacto or '—',
            p.email or '—', p.telefono or '—', n_prod,
        ], thin, alt=i%2==1)
        ws.cell(row=row, column=1).font = Font(bold=True, size=9)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer